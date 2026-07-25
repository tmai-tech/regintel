#!/usr/bin/env python3
"""Seed JSON catalog from BCI Tracking Plan.xlsx."""
from __future__ import annotations

import json
import shutil
from collections import OrderedDict
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "BCI Tracking Plan.xlsx"
DATA = ROOT / "data"
WEB_DATA = ROOT / "web" / "data"


def ser(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, float) and v == int(v):
        return int(v)
    if isinstance(v, str):
        return v.strip()
    return v


def sheet_to_records(ws, header_row=1):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = []
    for i, h in enumerate(rows[header_row - 1]):
        name = ser(h) or f"col_{i + 1}"
        headers.append(str(name).replace("\n", " ").strip())
    seen: dict[str, int] = {}
    clean = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            clean.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 1
            clean.append(h)
    records = []
    for row in rows[header_row:]:
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        rec = OrderedDict()
        for h, c in zip(clean, row):
            val = ser(c)
            if val is not None and val != "":
                rec[h] = val
        if rec:
            records.append(rec)
    return records


def main():
    if not XLSX.exists():
        raise SystemExit(f"Missing workbook: {XLSX}")
    wb = load_workbook(XLSX, data_only=True)
    DATA.mkdir(exist_ok=True)

    primary_raw = sheet_to_records(wb["Primary Links"])
    freq_keys = {k for r in primary_raw for k in r if "Frequency" in k or "frequency" in k}
    primary = []
    for r in primary_raw:
        topics = [str(v) for k, v in r.items() if ("Sub Category" in k or "Topic" in k) and v]
        freq = next((r[k] for k in freq_keys if k in r), None)
        url = r.get("Website links")
        status = "active"
        if not url:
            status = "missing_url"
        elif "not working" in str(url).lower():
            status = "broken"
        if r.get("Nature of weblink") and "not working" in str(r.get("Nature of weblink")).lower():
            status = "broken"
        primary.append(
            {
                "region": r.get("Region"),
                "jurisdiction": r.get("Jursidictions") or r.get("Jurisdictions"),
                "authority": r.get(
                    "Name of Regulatory Authority/ Concerned Ministry/ Government Website/ Concerned Department/ Secretary of State/Attorney General"
                ),
                "authority_type": r.get("Nature of Authority"),
                "link_nature": r.get("Nature of weblink"),
                "url": url,
                "frequency": freq,
                "segment": r.get("Segment (Private, Capital, Enterprise)"),
                "topics": topics,
                "status": status,
                "source_kind": "primary",
            }
        )

    tracking_raw = sheet_to_records(wb["Tracking Sheet"])
    tracking = []
    for r in tracking_raw:
        pub = r.get("Date of Publication on Subscribed Database (DD-MMM-YYYY)") or next(
            (r[k] for k in r if "Publication" in k), None
        )
        tracking.append(
            {
                "country": r.get("Country"),
                "federal_or_state": r.get("Federal/State"),
                "period_of_tracking": r.get("Period of Tracking"),
                "date_of_tracking": r.get("Date of Tracking"),
                "date_of_publication": pub,
                "law_area": r.get("Law Area"),
                "topical_relevance": r.get("Topical Relevance"),
                "link": r.get("Link for update"),
                "remarks": r.get("Remarks/Reasons for tracking"),
                "tracked_by": r.get("Tracked by"),
                "relevancy": r.get("Relevancy"),
                "comments": r.get("Comments"),
                "cor_impact": r.get("COR Impact (Yes/No)"),
                "assigned_to": r.get("Assigned to"),
                "slr_name": r.get("SLR Name"),
                "alert_status": r.get("Alert Status") or "new",
                "last_tracked": r.get("Last tracked"),
                "kmp_id": r.get("KMP ID"),
            }
        )

    gazette = []
    for r in sheet_to_records(wb["Gazette & Parliament Bills"]):
        gazette.append(
            {
                "jurisdiction": r.get("Jurisdiction"),
                "parliamentary_bills": r.get("Parliamentary Bills"),
                "official_gazette": r.get("Official Gazette/Legal Publications"),
                "legal_databases": r.get("Relevant Legal Databases/Sources"),
                "source_kind": "gazette",
            }
        )

    secondary = []
    for r in sheet_to_records(wb["Secondary Sources"]):
        vals = list(r.values())
        secondary.append(
            {
                "name": vals[0] if vals else None,
                "url": vals[1] if len(vals) > 1 else None,
                "coverage_area": vals[2] if len(vals) > 2 else None,
                "source_kind": "secondary",
                "status": "active",
            }
        )

    detailed = sheet_to_records(wb["Detailed Plan"])
    summary = sheet_to_records(wb["Summary Plan"])

    updates_path = DATA / "updates.json"
    updates = []
    if updates_path.exists():
        try:
            updates = json.loads(updates_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            updates = []

    meta = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "counts": {
            "primary_sources": len(primary),
            "tracking_records": len(tracking),
            "gazette_sources": len(gazette),
            "secondary_sources": len(secondary),
            "detailed_plan_rows": len(detailed),
            "updates": len(updates),
        },
    }

    payloads = {
        "meta.json": meta,
        "primary_sources.json": primary,
        "tracking.json": tracking,
        "gazette.json": gazette,
        "secondary_sources.json": secondary,
        "detailed_plan.json": detailed,
        "summary_plan.json": summary,
        "updates.json": updates,
    }
    for name, obj in payloads.items():
        (DATA / name).write_text(json.dumps(obj, indent=2, ensure_ascii=False), encoding="utf-8")

    WEB_DATA.mkdir(parents=True, exist_ok=True)
    for name in payloads:
        shutil.copy2(DATA / name, WEB_DATA / name)

    print(json.dumps(meta, indent=2))
    print(f"Wrote {len(payloads)} files to {DATA} and {WEB_DATA}")


if __name__ == "__main__":
    main()
